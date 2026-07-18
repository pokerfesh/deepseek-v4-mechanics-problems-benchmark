import csv, json, os, re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

out = os.path.expanduser("~/benchmarks")

# ── Collect best answers ──
all_runs = [f"ch{i}_ungrounded.csv" for i in range(1, 15)]
all_runs += ["supplement_ch1.csv", "supplement_ch4.csv", "supplement_ch5.csv",
             "rerun_empty_responses.csv", "rerun_persistent.csv",
             "rerun_all_empties.csv", "rerun_final_v2.csv",
             "rerun_multipart_1-3.csv", "rerun_multipart_4-6.csv",
             "rerun_salvage_empties.csv"]

best = {}
for fname in all_runs:
    path = os.path.join(out, fname)
    if not os.path.exists(path): continue
    with open(path) as f:
        next(f)
        for line in f:
            parts = line.split(",")
            if len(parts) < 8: continue
            prob = parts[2]; cond = parts[3]
            ans = parts[7].strip().strip('"')
            model = "Pro" if "Pro" in cond else "Flash"
            key = (prob, model)
            if ans and (key not in best or len(ans) > len(best[key])):
                best[key] = ans
            elif not ans and key not in best:
                best[key] = ""

with open(os.path.join(out, "university_physics_problems.json")) as f:
    answers = {p["problem"]: p["answer"] for p in json.load(f)}

# Group rows by chapter, only both-parsed
by_ch = defaultdict(list)
all_problems = sorted(set(k[0] for k in best), key=lambda x: (int(x.split(".")[0]), int(x.split(".")[1])))

chapter_names = {
    1: "Units & Vectors", 2: "Motion Along a Line", 3: "Motion in 2D 3D",
    4: "Newtons Laws", 5: "Applying Newtons Laws", 6: "Work & Kinetic Energy",
    7: "Potential Energy", 8: "Momentum & Collisions", 9: "Rotation of Rigid Bodies",
    10: "Dynamics of Rotation", 11: "Equilibrium & Elasticity", 12: "Fluid Mechanics",
    13: "Gravitation", 14: "Periodic Motion"
}

for pid in all_problems:
    ch = int(pid.split(".")[0])
    if ch > 14: continue
    p_ans = best.get((pid, "Pro"), "")
    f_ans = best.get((pid, "Flash"), "")
    if not p_ans or not f_ans: continue
    p_clean = re.sub(r'^FINAL ANSWER:\s*', '', p_ans).strip()
    f_clean = re.sub(r'^FINAL ANSWER:\s*', '', f_ans).strip()
    book = answers.get(pid, "")
    by_ch[ch].append((ch, pid, book, p_clean, "", f_clean, "", ""))

# ── Write XLSX ──
xlsx_path = os.path.join(out, "mechanics_ch1-14_clean.xlsx")
wb = Workbook()
wb.remove(wb.active)

header = ["Chapter", "Problem #", "Answer (Book)", "Pro Response",
          "Pro Correct?", "Flash Response", "Flash Correct?", "Remarks"]
hfont = Font(bold=True, size=10, color="FFFFFF")
hfill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
afill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
widths = [9, 11, 42, 50, 14, 50, 14, 22]

all_rows = []

for ch in range(1, 15):
    rows = sorted(by_ch[ch], key=lambda r: int(r[1].split(".")[1]))
    all_rows.extend(rows)
    
    tab_name = f"Ch.{ch} {chapter_names.get(ch, '')}"[:31]
    ws = wb.create_sheet(title=tab_name)
    
    for ci, h in enumerate(header, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    
    for ri, r in enumerate(rows, 2):
        for ci, val in enumerate(r, 1):
            c = ws.cell(row=ri, column=ci, value=val if val else "")
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if ci == 3: c.fill = afill
    
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

# Summary tab
ws = wb.create_sheet(title="Summary", index=0)
sfill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
sfont = Font(bold=True, size=12, color="FFFFFF")
h2font = Font(bold=True, size=11)

ws.cell(row=1, column=1, value="BENCHMARK SUMMARY (Both Models Parsed)").font = sfont
ws.cell(row=1, column=1).fill = sfill
ws.merge_cells("A1:E1")

row_num = 3
ws.cell(row=row_num, column=1, value="Chapter").font = h2font
ws.cell(row=row_num, column=2, value="Problems").font = h2font
ws.cell(row=row_num, column=3, value="Pro Correct").font = h2font
ws.cell(row=row_num, column=4, value="Flash Correct").font = h2font
ws.cell(row=row_num, column=5, value="Both Correct").font = h2font

row_num += 1
tab_total = 0
pro_correct_total = 0
flash_correct_total = 0
both_correct_total = 0

for ch in range(1, 15):
    rows = by_ch[ch]
    tab_name = f"Ch.{ch}"
    n = len(rows)
    tab_total += n
    
    ws.cell(row=row_num, column=1, value=f"Ch.{ch}")
    ws.cell(row=row_num, column=2, value=n)
    
    # Formulas referencing each chapter tab
    pro_formula = f"=COUNTIF('{tab_name}'!E:E,TRUE)"
    flash_formula = f"=COUNTIF('{tab_name}'!G:G,TRUE)"
    both_formula = f"=COUNTIFS('{tab_name}'!E:E,TRUE,'{tab_name}'!G:G,TRUE)"
    
    ws.cell(row=row_num, column=3, value=pro_formula)
    ws.cell(row=row_num, column=4, value=flash_formula)
    ws.cell(row=row_num, column=5, value=both_formula)
    
    row_num += 1

row_num += 1
ws.cell(row=row_num, column=1, value="TOTAL").font = h2font
ws.cell(row=row_num, column=2, value=tab_total)
pro_tot_rng = f"C5:C18"
flash_tot_rng = f"D5:D18"
both_tot_rng = f"E5:E18"

ws.cell(row=row_num, column=3, value=f"=SUM({pro_tot_rng})")
ws.cell(row=row_num, column=4, value=f"=SUM({flash_tot_rng})")
ws.cell(row=row_num, column=5, value=f"=SUM({both_tot_rng})")

row_num += 1
ws.cell(row=row_num, column=1, value="Pro Accuracy").font = h2font
ws.cell(row=row_num, column=2, value=f'=IF(SUM({pro_tot_rng})+SUM(C{6}:C{18})-SUM({pro_tot_rng})=0,"",{pro_tot_rng}/B{row_num-1})')
# Simpler: just compute total accuracy
ws.cell(row=row_num, column=2, value=f'=IF(B{row_num-1}>0,SUM({pro_tot_rng})/B{row_num-1},"")')
ws.cell(row=row_num, column=2).number_format = "0.0%"

ws.cell(row=row_num+1, column=1, value="Flash Accuracy").font = h2font
ws.cell(row=row_num+1, column=2, value=f'=IF(B{row_num-1}>0,SUM({flash_tot_rng})/B{row_num-1},"")')
ws.cell(row=row_num+1, column=2).number_format = "0.0%"

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 15
ws.column_dimensions["E"].width = 14

wb.save(xlsx_path)

# Also write CSV
csv_path = os.path.join(out, "mechanics_ch1-14_clean.csv")
with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["Chapter", "Problem #", "Answer (Book)", "Pro Response",
                "Pro Correct?", "Flash Response", "Flash Correct?", "Remarks"])
    for r in all_rows:
        w.writerow(list(r))

total_all = len(all_rows)
print(f"XLSX: {xlsx_path}")
print(f"CSV:  {csv_path}")
print(f"\n{total_all} problems, 14 chapter tabs + Summary tab")
print()
print("To use checkboxes in Google Sheets:")
print("  1. Upload XLSX to sheets.google.com")
print("  2. In each chapter tab, select cells E2:E? (Pro Correct?)")
print("  3. Hold Ctrl, select G2:G? (Flash Correct?)")
print("  4. Menu: Insert > Checkbox")
print("  5. Summary tab auto-counts checkboxes across all tabs")
